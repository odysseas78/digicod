import pandas
from openpyxl import load_workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font
from openpyxl.worksheet.hyperlink import Hyperlink




# cqs = Order.objects.filter(created_at__gt='2022-06-01',created_at__lt='2024-03-31', status='completed')
# # logstatqs = LoginStatistic.objects.filter(result=891045).first()
# # print(logstatqs)

# a=0
# lst=[]
# for f in cqs:
#     invnr = str(f.created_at)[:10].replace('-','')[2:]+str(f.id)
#     logstatqs = LoginStatistic.objects.filter(ip=f.ip).first()
#     country = logstatqs.geopos if logstatqs else logstatqs
#     print(f'''Date: {str(f.created_at)[:19]} - Order-Nr.: {f.id} - Invoice-Nr.: DC-{invnr} - Total: {float(f.pay_amount)} Currency: {f.pay_currency.shortname} -\
# Total EUR: {float(f.europrice)} - Customer: {f.customer.user.email} Country: {country}''')
#     obj={
#         'Date':(str(f.created_at)[:19]),
#         'Order-Nr':f.id,
#         'Invoice-Nr': f'DC-{invnr}',
#         # 'Total': float(f.pay_amount),
#         # 'Payment currency': f.pay_currency.shortname,
#         'Total EUR': float(f.europrice),
#         'Customer': f.customer.user.email,
#         'Country': country
        
#     }
#     lst.append(obj)
#     # a+=1
#     # if a == 225:
#     #     break
# with open('invoices.json', 'w') as f:
#     f.write(json.dumps(lst))
   
# pandas.read_json('invoices.json').to_excel('invoices.xlsx')

# bqs = Brand.objects.filter(in_stock=True, active=True, deleted=False).exclude(category__slug='region-free')
# cqs = Category.objects.exclude(slug='region-free')
# # pqs = Product.objects.filter(in_stock=True, active=True, deleted=False, qty__gt=0).exclude(brand__category__slug='region-free')
# # logstatqs = LoginStatistic.objects.filter(result=891045).first()
# # print(bqs.first().products.all().filter(in_stock=True, active=True, deleted=False, qty__gt=0))
# # print(pqs)


# a=0
# lst=[]
# for c in cqs:
#     for f in bqs.filter(category__slug=c.slug):
#         if f.products.all().filter(in_stock=True, active=True, deleted=False, qty__gt=0).count() > 0:
#             print(f'''Title: {f.title} - Category: {f.category.all().first().name} - url: https://digicod.eu/products/{f.category.all().first().slug}/{f.slug}''')
#             obj={
#                 'Title':(f.title),
#                 'category':f.category.all().first().name,
#                 'Url': f'https://digicod.eu/products/{f.category.all().first().slug}/{f.slug}',

#             }
#             lst.append(obj)
#             a+=1
#     # if a == 225:
#     #     break
# print(a)
# with open('product_brands.json', 'w') as f:
#     f.write(json.dumps(lst))
   
# pandas.read_json('product_brands.json').to_excel('product_brands.xlsx')



# wb = load_workbook('product_brands.xlsx')
# ws = wb.active

# for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=4):
#     cell = row[3]  # Assuming 'Url' is in the third column
#     url = cell.value
#     cell.hyperlink = url
#     cell.value = "Link"
#     cell.font = Font(color="0000FF", underline="single")

# # Datei speichern
# wb.save('product_brands.xlsx')